"""
construction_agent.py
=======================

This module contains a simplified prototype of a construction manager agent.
It demonstrates how to parse a colour‑coded Excel schedule, infer activity
durations, compute a basic critical‑path schedule, estimate labour costs and
contingency allowances, and produce a summary report via the console.

The design is deliberately modular: each step (parsing, dependency inference,
CPM calculation, cost estimation) is encapsulated in a function.  This makes
it easier to extend individual pieces (for example, adding more sophisticated
dependency heuristics) without rewriting the entire pipeline.

The script can be invoked from the command line:

    python construction_agent.py --file "B13 Remodeling Schedule.xlsx" \
        --start-date "2025-08-01" --hours-per-day 8 \
        --base-rate 80 --labour-burden 0.2 --inefficiency 0.2 \
        --contingency 0.07

The above example assumes Day 1 corresponds to 1 August 2025, uses an 8‑hour
work day, a crew base rate of $80 per hour, a 20 % labour‑burden rate, a
20 % inefficiency factor and a 7 % contingency for renovations.  These
parameters can be adjusted to suit local conditions.

Limitations: The dependency graph is currently a simple sequential chain.
Real projects may have complex parallel activities.  You should supply
dependencies explicitly or implement heuristics to deduce them.
"""

from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import openpyxl


@dataclass
class Activity:
    """Represents a construction activity extracted from the schedule."""
    name: str
    start_day: int
    end_day: int
    duration: int = field(init=False)
    start_date: Optional[dt.date] = None
    end_date: Optional[dt.date] = None
    es: Optional[int] = None  # earliest start (day number)
    ef: Optional[int] = None  # earliest finish (day number)
    ls: Optional[int] = None  # latest start (day number)
    lf: Optional[int] = None  # latest finish (day number)
    slack: Optional[int] = None
    labour_hours: Optional[float] = None
    labour_cost: Optional[float] = None

    def __post_init__(self) -> None:
        self.duration = self.end_day - self.start_day + 1


def parse_schedule(file_path: str) -> List[Activity]:
    """
    Parse a colour‑coded Excel schedule and return a list of activity objects.

    The function expects that the second row of the sheet contains day labels
    (e.g., "Day 1", "Day 2", …).  Coloured cells or cells containing text
    indicate that work is scheduled on that day.  Consecutive coloured cells
    are grouped into activity segments.

    Parameters
    ----------
    file_path : str
        Path to the Excel file.

    Returns
    -------
    List[Activity]
        A list of activities with start and end day indices.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active

    # Identify header row containing day labels.  We assume day labels start
    # with "Day" (e.g., "Day 1").  We'll search row 2 by default but we
    # also fall back to scanning the first few rows.
    day_row_idx = None
    for r in range(1, min(10, ws.max_row + 1)):
        cells = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(isinstance(v, str) and v.startswith("Day ") for v in cells):
            day_row_idx = r
            break
    if day_row_idx is None:
        raise ValueError("Could not locate a row with day labels (e.g., 'Day 1').")

    # Map column indices to day numbers.
    day_cols: List[Tuple[int, int]] = []
    for col in range(2, ws.max_column + 1):
        val = ws.cell(row=day_row_idx, column=col).value
        if isinstance(val, str) and val.startswith("Day "):
            try:
                day_num = int(val.split()[1])
                day_cols.append((col, day_num))
            except (IndexError, ValueError):
                continue

    if not day_cols:
        raise ValueError("No day columns found in the schedule. Check the header row.")

    activities: List[Activity] = []
    # Start parsing rows below the header row.  We skip empty names.
    for row in range(day_row_idx + 1, ws.max_row + 1):
        name_cell = ws.cell(row=row, column=1).value
        if not name_cell or str(name_cell).strip() == "":
            continue
        name = str(name_cell).strip()
        in_segment = False
        seg_start_day: Optional[int] = None
        for col, day in day_cols:
            cell = ws.cell(row=row, column=col)
            # Determine if the cell is coloured or contains text.  openpyxl stores
            # fill colours using patternType and fgColor; we check patternType.
            coloured = False
            if cell.fill and cell.fill.patternType:
                # For solid fills, patternType is 'solid'.  We also check for
                # foreground colour type 'rgb'.  Some Excel cells may use
                # indexed colours; these are considered coloured as well.
                if cell.fill.patternType != "none":
                    coloured = True
            # Consider a cell with a value as indicating work as well.
            has_value = cell.value not in (None, "", 0)

            if coloured or has_value:
                if not in_segment:
                    in_segment = True
                    seg_start_day = day
            else:
                if in_segment:
                    # End of a segment.
                    if seg_start_day is None:
                        raise RuntimeError("Unexpected state: segment started without a start day.")
                    activities.append(Activity(name=name, start_day=seg_start_day, end_day=day - 1))
                    in_segment = False
                    seg_start_day = None
        # If segment continues until the last day column, close it.
        if in_segment and seg_start_day is not None:
            last_day = day_cols[-1][1]
            activities.append(Activity(name=name, start_day=seg_start_day, end_day=last_day))

    return activities


def assign_dates(activities: List[Activity], start_date: dt.date) -> None:
    """
    Assign calendar dates to each activity using the reference start date.

    The first day in the schedule (Day 1) is mapped to `start_date`.  Each
    activity's start_day and end_day are converted to dates accordingly.

    Parameters
    ----------
    activities : List[Activity]
        Parsed activities with start_day and end_day fields.
    start_date : datetime.date
        Calendar date corresponding to Day 1.
    """
    for activity in activities:
        activity.start_date = start_date + dt.timedelta(days=activity.start_day - 1)
        activity.end_date = start_date + dt.timedelta(days=activity.end_day - 1)


def infer_dependencies(activities: List[Activity]) -> Dict[str, List[str]]:
    """
    Infer simple sequential dependencies for activities.

    This function orders activities by their start_day and, if multiple
    activities share the same start_day, by their names.  Each activity
    (except the first) is assumed to depend on the immediately preceding
    activity.  While simplistic, this ensures a deterministic DAG for CPM
    computation.  Users can replace or extend this function with more
    sophisticated heuristics or by reading dependencies from the schedule.

    Returns
    -------
    Dict[str, List[str]]
        A mapping from each activity name to a list of its predecessor names.
    """
    sorted_acts = sorted(activities, key=lambda a: (a.start_day, a.name))
    deps: Dict[str, List[str]] = {act.name: [] for act in sorted_acts}
    for idx, act in enumerate(sorted_acts):
        if idx > 0:
            prev = sorted_acts[idx - 1]
            # Avoid self‑dependency in case of duplicate names at different rows.
            if act.name != prev.name:
                deps[act.name].append(prev.name)
    return deps


def compute_cpm(activities: List[Activity], dependencies: Optional[Dict[str, List[str]]] = None) -> None:
    """
    Compute CPM scheduling parameters (ES, EF, LS, LF, slack) for each activity.

    This simplified implementation assumes a sequential dependency structure: the
    activities are sorted by their start_day (and name as a tie breaker) and
    executed one after another.  Each activity (except the first) depends on
    the previous activity.  The earliest start time (ES) of the first
    activity is day 1.  For subsequent activities, ES = previous EF + 1.

    Latest finish times (LF) are computed by traversing the activities in
    reverse order: the last activity must finish by the project end (the
    maximum EF), and each preceding activity must finish just before the
    successor starts.  Slack is LS - ES; with sequential dependencies, all
    activities will typically have zero slack (critical path).

    Parameters
    ----------
    activities : List[Activity]
        Activities to schedule.  Durations must be computed before calling
        this function.
    dependencies : Optional[Dict[str, List[str]]], optional
        Unused in this simplified version; retained for API compatibility.
    """
    # Sort activities by start_day and name to ensure deterministic order.
    sorted_acts = sorted(activities, key=lambda a: (a.start_day, a.name))

    # Forward pass: compute earliest start (ES) and finish (EF) times.
    prev_ef = None
    for idx, act in enumerate(sorted_acts):
        if idx == 0:
            act.es = 1
        else:
            # Start immediately after predecessor finishes.
            act.es = prev_ef + 1 if prev_ef is not None else 1
        act.ef = act.es + act.duration - 1
        prev_ef = act.ef

    # The end of the project is the EF of the last activity.
    project_end = prev_ef or 0

    # Backward pass: compute latest finish (LF) and start (LS) times.
    next_ls = None
    for idx, act in enumerate(reversed(sorted_acts)):
        if idx == 0:
            # Last activity finishes at project end.
            act.lf = project_end
        else:
            # Finish just before the successor starts.
            act.lf = next_ls - 1 if next_ls is not None else project_end
        act.ls = act.lf - act.duration + 1
        act.slack = (act.ls - act.es) if act.es is not None else None
        next_ls = act.ls


def estimate_costs(
    activities: List[Activity],
    hours_per_day: float = 8.0,
    base_rate: float = 80.0,
    labour_burden: float = 0.2,
    inefficiency: float = 0.2,
) -> Tuple[float, float]:
    """
    Estimate labour hours and costs for each activity and return totals.

    Parameters
    ----------
    activities : List[Activity]
        Activities with durations defined.
    hours_per_day : float, optional
        Number of working hours per day (default: 8).  This should not
        exceed statutory limits; Indian law caps daily working hours at nine【886789385528665†L2169-L2174】.
    base_rate : float, optional
        Base hourly rate for the crew (default: $80).  Users should supply
        realistic values based on local wages and crew composition【143656559969249†L229-L243】.
    labour_burden : float, optional
        Fractional labour burden (e.g., 0.2 for 20 %) covering payroll
        taxes, insurance and statutory contributions【143656559969249†L250-L284】.
    inefficiency : float, optional
        Fractional inefficiency factor (e.g., 0.2 for 20 %) accounting for
        unproductive time【143656559969249†L286-L312】.

    Returns
    -------
    Tuple[float, float]
        Total labour hours and total labour cost across all activities.
    """
    total_hours: float = 0.0
    total_cost: float = 0.0
    hourly_rate = base_rate * (1 + labour_burden + inefficiency)
    for act in activities:
        act.labour_hours = act.duration * hours_per_day
        total_hours += act.labour_hours
        act.labour_cost = act.labour_hours * hourly_rate
        total_cost += act.labour_cost
    return total_hours, total_cost


def compute_contingency(base_cost: float, contingency_rate: float = 0.07) -> float:
    """
    Compute contingency allowance based on base cost.

    A renovation project typically carries a contingency of 7–8 %【537146304403148†L190-L209】.

    Parameters
    ----------
    base_cost : float
        Total labour and material cost prior to contingency.
    contingency_rate : float, optional
        Fractional rate (e.g., 0.07 for 7 %).

    Returns
    -------
    float
        Contingency amount.
    """
    return base_cost * contingency_rate


def print_summary(
    activities: List[Activity],
    total_hours: float,
    total_cost: float,
    contingency_rate: float = 0.07,
) -> None:
    """
    Print a console summary of the schedule, costs and CPM analysis.

    Activities are sorted by earliest start day for readability.  The summary
    includes each activity's start and finish dates, duration, labour hours
    and cost, plus whether it is on the critical path (slack = 0).

    At the end, the function prints total labour hours, base cost, contingency
    allowance and overall cost including contingency.
    """
    print("\n=== Activity Schedule and Cost Summary ===\n")
    sorted_acts = sorted(activities, key=lambda a: (a.es or 0, a.name))
    header = [
        "Activity", "Start", "Finish", "Duration (d)",
        "Labour hours", "Cost ($)", "Critical?"
    ]
    print("{:<25} {:<12} {:<12} {:<12} {:<14} {:<12} {}".format(*header))
    print("-" * 100)
    for act in sorted_acts:
        start = act.start_date.strftime("%Y-%m-%d") if act.start_date else f"Day {act.start_day}"
        finish = act.end_date.strftime("%Y-%m-%d") if act.end_date else f"Day {act.end_day}"
        crit = "YES" if act.slack == 0 else "no"
        print(
            f"{act.name:<25} {start:<12} {finish:<12} "
            f"{act.duration:<12} {act.labour_hours:<14.1f} {act.labour_cost:<12.2f} {crit}"
        )
    contingency_amount = compute_contingency(total_cost, contingency_rate)
    total_with_contingency = total_cost + contingency_amount
    print("\nTotals:")
    print(f"  Total labour hours: {total_hours:.1f} h")
    print(f"  Total labour cost: ${total_cost:,.2f}")
    print(f"  Contingency ({contingency_rate*100:.0f}%): ${contingency_amount:,.2f}")
    print(f"  Total including contingency: ${total_with_contingency:,.2f}\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="AI‑powered construction manager prototype")
    parser.add_argument("--file", required=True, help="Path to the Excel schedule (colour‑coded)")
    parser.add_argument("--start-date", required=True, help="Calendar date for Day 1 (YYYY-MM-DD)")
    parser.add_argument("--hours-per-day", type=float, default=8.0, help="Number of working hours per day (<=9)")
    parser.add_argument("--base-rate", type=float, default=80.0, help="Base hourly labour rate ($)")
    parser.add_argument("--labour-burden", type=float, default=0.2, help="Labour burden fraction (e.g., 0.2 for 20%)")
    parser.add_argument("--inefficiency", type=float, default=0.2, help="Inefficiency factor fraction")
    parser.add_argument("--contingency", type=float, default=0.07, help="Contingency fraction (e.g., 0.07 for 7%)")
    args = parser.parse_args()

    # Parse activities from schedule
    activities = parse_schedule(args.file)
    if not activities:
        print("No activities found in the provided schedule.")
        return

    # Assign calendar dates to activities
    start_date = dt.datetime.strptime(args.start_date, "%Y-%m-%d").date()
    assign_dates(activities, start_date)

    # Infer dependencies (simple sequential chain)
    deps = infer_dependencies(activities)

    # Compute CPM schedule
    compute_cpm(activities, deps)

    # Update calendar dates based on CPM results (earliest start/finish)
    for act in activities:
        if act.es is not None:
            act.start_date = start_date + dt.timedelta(days=act.es - 1)
        if act.ef is not None:
            act.end_date = start_date + dt.timedelta(days=act.ef - 1)

    # Estimate labour costs
    total_hours, total_cost = estimate_costs(
        activities,
        hours_per_day=args.hours_per_day,
        base_rate=args.base_rate,
        labour_burden=args.labour_burden,
        inefficiency=args.inefficiency,
    )

    # Print summary report
    print_summary(
        activities,
        total_hours,
        total_cost,
        contingency_rate=args.contingency,
    )


if __name__ == "__main__":
    main()
